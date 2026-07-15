#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Fluxo de Caixa Mensal — Loteamento Papagaios/MG
Arquitetura: Parâmetros → Cronograma (fórmulas) → Abas anuais (referências)
Altere Parâmetros ou Cronograma e tudo recalcula automaticamente.
"""

import sys
sys.path.insert(0, '/Users/leonardoforx/Library/Python/3.9/lib/python/site-packages')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
wb.remove(wb.active)

# ── CORES ──────────────────────────────────
AZUL_ESC="1F4E79"; AZUL_MED="2E75B6"; AZUL_CLA="D6E4F0"
VERDE_ESC="1A5276"; VERDE_CLA="D5E8D4"
CINZA="F2F2F2"; BRANCO="FFFFFF"; PRETO="000000"
AMARELO="FFF2CC"; LARANJA="FCE4D6"
VERDE_SEMAF="C8F0D0"; VERM_SEMAF="FFBCBC"

def fill(c): return PatternFill("solid", fgColor=c)
def bord():
    s = Side(style='thin', color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)
def bord_med():
    s = Side(style='medium', color="1F4E79")
    return Border(left=s, right=s, top=s, bottom=s)

def cel(ws, row, col, value=None, bg=BRANCO, bold=False, size=9,
        color=PRETO, fmt=None, h='right', wrap=False, thick=False, italic=False):
    c = ws.cell(row=row, column=col)
    if value is not None: c.value = value
    c.fill = fill(bg)
    c.font = Font(bold=bold, size=size, color=color, name='Calibri', italic=italic)
    c.alignment = Alignment(horizontal=h, vertical='center', wrap_text=wrap)
    c.border = bord_med() if thick else bord()
    if fmt: c.number_format = fmt
    return c

MOEDA='R$ #,##0'; MOEDA2='R$ #,##0.00'; INT='#,##0'; PERC='0.0%'

MESES = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
         "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]

# ══════════════════════════════════════════════════════
# ABA 1 — PARÂMETROS (única aba com valores diretos)
# ══════════════════════════════════════════════════════
wsp = wb.create_sheet("Parâmetros")
wsp.column_dimensions['A'].width = 42
wsp.column_dimensions['B'].width = 22
wsp.column_dimensions['C'].width = 14
wsp.column_dimensions['D'].width = 35

wsp.merge_cells('A1:D1')
c = wsp['A1']; c.value = "PARÂMETROS — altere aqui e todo o fluxo de caixa recalcula"
c.fill = fill(AZUL_ESC); c.font = Font(bold=True, size=12, color=BRANCO, name='Calibri')
c.alignment = Alignment(horizontal='center', vertical='center')
wsp.row_dimensions[1].height = 26

wsp.merge_cells('A2:D2')
c = wsp['A2']
c.value = "🟡 Amarelo = editável   |   🔵 Azul = calculado automaticamente"
c.fill = fill(AMARELO); c.font = Font(size=9, italic=True, color="7D6608", name='Calibri')
c.alignment = Alignment(horizontal='left', vertical='center')
wsp.row_dimensions[2].height = 15

def par(ws, row, nome, val, unid="", obs="", editable=True, fmt=None):
    ws.row_dimensions[row].height = 17
    bg_n = CINZA if row%2==0 else BRANCO
    bg_v = AMARELO if editable else AZUL_CLA
    cel(ws, row, 1, nome, bg=bg_n, h='left', size=10)
    cel(ws, row, 2, val,  bg=bg_v, bold=editable, h='right', size=10, fmt=fmt)
    cel(ws, row, 3, unid, bg=bg_n, italic=True, color="595959", h='left', size=9)
    cel(ws, row, 4, obs,  bg=bg_n, color="595959", h='left', size=9)

def par_sec(ws, row, titulo):
    ws.merge_cells(f'A{row}:D{row}')
    c = ws.cell(row=row, column=1, value=titulo)
    c.fill = fill(AZUL_ESC); c.font = Font(bold=True, size=10, color=BRANCO, name='Calibri')
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = bord(); ws.row_dimensions[row].height = 18

par_sec(wsp, 3, "COMERCIAL")
# P_PRECO = B4, P_LOTES = B5, ...
par(wsp, 4,  "Preço de venda por lote (R$)",    80000,  "R$",      "Entrada — loteamento popular",    fmt=MOEDA)
par(wsp, 5,  "Total de lotes",                  330,    "un.",     "Aprovado pelo projeto urbanístico")
par(wsp, 6,  "VGV Total (R$)",                  "=B4*B5","R$",     "Calculado", editable=False, fmt=MOEDA)

par_sec(wsp, 7, "FINANCIAMENTO")
par(wsp, 8,  "% lotes à vista / banco",         0.70,   "",        "70% recebido na venda",           fmt=PERC)
par(wsp, 9,  "% lotes financiados empreendedor","=1-B8","",        "Calculado", editable=False,        fmt=PERC)
par(wsp, 10, "Prazo financiamento (meses)",      60,    "meses",   "Máximo 60 meses")
par(wsp, 11, "Parcela por lote (R$/mês)",        "=B4*B9/B10","R$/lote/mês","Calculado",editable=False,fmt=MOEDA2)

par_sec(wsp, 12, "TAXAS E PERCENTUAIS")
par(wsp, 13, "Taxa de corretagem",               0.06,  "% venda", "6% sobre valor de venda",         fmt=PERC)
par(wsp, 14, "Taxa administração (% VGV)",       0.03,  "% VGV",   "3% do VGV total",                 fmt=PERC)
par(wsp, 15, "Taxa marketing (% VGV)",           0.02,  "% VGV",   "2% do VGV total",                 fmt=PERC)
par(wsp, 16, "Taxa contingência",                0.03,  "% custos","3% sobre custos diretos",          fmt=PERC)
par(wsp, 17, "Taxa impostos — Lucro Presumido",  0.065, "% receita","PIS+COFINS+IRPJ+CSLL",           fmt=PERC)
par(wsp, 18, "% Empreendedor",                   0.65,  "",        "Base: lucro antes dos impostos",   fmt=PERC)
par(wsp, 19, "% Terrenista",                     "=1-B18","",     "Calculado", editable=False,         fmt=PERC)

par_sec(wsp, 20, "CAPITAL")
par(wsp, 21, "Capital próprio investido (R$)",   1000000,"R$",     "Aporte inicial — mês 0",           fmt=MOEDA)
par(wsp, 22, "Taxa custo financeiro (a.m.)",     0.015, "a.m.",    "1,5% sobre saldo negativo acum.",  fmt=PERC)

par_sec(wsp, 23, "INFRAESTRUTURA")
par(wsp, 24, "Total infraestrutura (R$)",        5430000,"R$",     "Da aba Infraestrutura — pode vincular se quiser",fmt=MOEDA)

wsp.merge_cells('A26:D26')
c = wsp['A26']
c.value = ("💡  DICA: Para vincular o total da infra automaticamente, substitua o valor de B24 pela "
           "fórmula: =Viabilidade!Infraestrutura!E14  (após gerar a planilha principal)")
c.fill = fill(AZUL_CLA); c.font = Font(size=9, italic=True, color=AZUL_ESC, name='Calibri')
c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
c.border = bord(); wsp.row_dimensions[26].height = 30

# Nomes de referência para facilitar leitura das fórmulas
# P4=preço, P5=lotes, P6=VGV, P8=%vista, P9=%financ, P10=prazo,
# P11=parcela, P13=corret, P14=admin, P15=mkt, P16=conting,
# P17=impostos, P18=%emp, P19=%terr, P21=capital, P22=taxa_fin, P24=infra

def P(campo):
    """Retorna referência absoluta à célula de parâmetro."""
    return {
        'preco': "Parâmetros!$B$4",
        'lotes': "Parâmetros!$B$5",
        'vgv':   "Parâmetros!$B$6",
        'vista': "Parâmetros!$B$8",
        'financ':"Parâmetros!$B$9",
        'prazo': "Parâmetros!$B$10",
        'parcela':"Parâmetros!$B$11",
        'corret':"Parâmetros!$B$13",
        'admin': "Parâmetros!$B$14",
        'mkt':   "Parâmetros!$B$15",
        'conting':"Parâmetros!$B$16",
        'impost':"Parâmetros!$B$17",
        'emp':   "Parâmetros!$B$18",
        'terr':  "Parâmetros!$B$19",
        'capital':"Parâmetros!$B$21",
        'taxa_fin':"Parâmetros!$B$22",
        'infra': "Parâmetros!$B$24",
    }[campo]

# ══════════════════════════════════════════════════════
# ABA 2 — CRONOGRAMA (motor de cálculo — 48 meses)
# ══════════════════════════════════════════════════════
# Colunas:
# A:Mês  B:Ano  C:Nome  D:Lotes(input)  E:Infra(input)  F:Mkt(input)
# G:Receita vista  H:Parcelas  I:Total Receita
# J:Corretagem  K:Admin  L:Conting  M:Total Custos
# N:Res.Bruto  O:Terrenista  P:Res.Empreendedor
# Q:Impostos  R:Custo Financeiro  S:Fluxo Líq.  T:Acumulado

wsc = wb.create_sheet("Cronograma")

# Larguras
largs = [8,6,13,10,16,14,16,18,16,13,14,13,15,16,15,16,13,16,16,18]
cols_letters = [get_column_letter(i+1) for i in range(len(largs))]
for l, larg in zip(cols_letters, largs):
    wsc.column_dimensions[l].width = larg

# Título
wsc.merge_cells(f'A1:{cols_letters[-1]}1')
c = wsc['A1']
c.value = "CRONOGRAMA MESTRE — motor de cálculo — todas as fórmulas ligadas a Parâmetros"
c.fill = fill(AZUL_ESC); c.font = Font(bold=True, size=11, color=BRANCO, name='Calibri')
c.alignment = Alignment(horizontal='center', vertical='center')
wsc.row_dimensions[1].height = 22

# Grupos de colunas (linha 2)
grupos = [("A:C","IDENTIFICAÇÃO",AZUL_ESC), ("D:F","ENTRADAS (EDITÁVEIS)",VERDE_ESC),
          ("G:I","RECEITAS","117A65"), ("J:M","CUSTOS","C0392B"),
          ("N:P","RESULTADO BRUTO",AZUL_ESC), ("Q:T","RESULTADO LÍQUIDO EMPREENDEDOR","1A5276")]
for rng, titulo, bg in grupos:
    s, e = rng.split(":")
    sc, ec = wsc.column_dimensions[s].width, wsc.column_dimensions[e].width
    cs = wsc[f'{s}2']; ce = wsc[f'{e}2']
    wsc.merge_cells(f'{s}2:{e}2')
    c = wsc[f'{s}2']; c.value = titulo
    c.fill = fill(bg); c.font = Font(bold=True, size=8, color=BRANCO, name='Calibri')
    c.alignment = Alignment(horizontal='center', vertical='center'); c.border = bord()
wsc.row_dimensions[2].height = 14

# Cabeçalho de colunas (linha 3)
hdrs = ["MÊS","ANO","MÊS\nNOME","LOTES\nVENDIDOS","INFRA-\nESTRUTURA","MARKETING\nMENSAL",
        "RECEITA\nÀ VISTA","PARCELAS\nRECEB.","TOTAL\nRECEITA",
        "CORRETA-\nGEM","ADMIN\nMENSAL","CONTIN-\nGÊNCIA","TOTAL\nCUSTOS",
        "RESULTADO\nBRUTO","PAGTO.\nTERRENISTA","RES.EMPREEN-\nDEDOR",
        "IMPOSTOS","CUSTO\nFINANC.","FLUXO\nLÍQUIDO","FLUXO\nACUMULADO"]
for i, h in enumerate(hdrs):
    c = wsc.cell(row=3, column=i+1, value=h)
    c.fill = fill(AZUL_ESC if i < 3 else (VERDE_ESC if i < 6 else
             ("117A65" if i < 9 else ("C0392B" if i < 13 else
             (AZUL_ESC if i < 16 else "1A5276")))))
    c.font = Font(bold=True, size=8, color=BRANCO, name='Calibri')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = bord()
wsc.row_dimensions[3].height = 30

# ── Cronograma de vendas (lotes por mês) ──
def lotes_mes(m):
    if m < 7:   return 0
    if m == 7:  return 40
    if m <= 36: return 10
    return 0

# ── Cronograma de infraestrutura ──
INFRA_SCHED = {
     1: 50000,  2:100000,  3:100000,
     4:200000,  5:300000,  6:300000,
     7:400000,  8:400000,  9:400000,
    10:400000, 11:400000, 12:380000,
    13:350000, 14:350000, 15:300000,
    16:200000, 17:150000, 18:100000,
    19:100000, 20: 80000, 21: 60000,
    22: 50000, 23: 30000, 24: 30000,
}
infra_def = sum(INFRA_SCHED.values())
infra_rest = 5430000 - infra_def
for m in range(25, 49):
    INFRA_SCHED[m] = round(infra_rest / 24, 0)

# ── Marketing mensal ──
MKT_SCHED = {5:50000, 6:100000, 7:100000, 8:50000, 9:30000,
             10:20000, 11:20000, 12:20000}
for m in range(13, 25): MKT_SCHED[m] = 5000
for m in range(1, 5):   MKT_SCHED.setdefault(m, 0)
mkt_def = sum(MKT_SCHED.values())
mkt_rest = max(0, 528000 - mkt_def)
for m in range(25, 49): MKT_SCHED[m] = round(mkt_rest / 24, 0)

# ── Linhas de dados (meses 1-48 → linhas 4-51) ──
DATA_START = 4   # linha onde começa o mês 1

for m in range(1, 49):
    r = DATA_START + m - 1
    wsc.row_dimensions[r].height = 15
    bg = CINZA if m % 2 == 0 else BRANCO
    ano = (m - 1) // 12 + 1
    mes_no_ano = (m - 1) % 12

    # Col A: número do mês
    cel(wsc, r, 1, m, bg=bg, h='center', size=9)
    # Col B: ano
    cel(wsc, r, 2, f"=INT((A{r}-1)/12)+1", bg=bg, h='center', size=9, fmt=INT)
    # Col C: nome do mês
    cel(wsc, r, 3, MESES[mes_no_ano], bg=bg, h='left', size=9)

    # Col D: lotes vendidos — ENTRADA EDITÁVEL
    cel(wsc, r, 4, lotes_mes(m), bg=AMARELO if lotes_mes(m) > 0 else bg,
        bold=lotes_mes(m) > 0, h='center', size=9, fmt=INT)

    # Col E: infraestrutura — ENTRADA EDITÁVEL
    cel(wsc, r, 5, INFRA_SCHED.get(m, 0), bg=AMARELO, h='right', size=9, fmt=MOEDA)

    # Col F: marketing mensal — ENTRADA EDITÁVEL
    cel(wsc, r, 6, MKT_SCHED.get(m, 0), bg=AMARELO, h='right', size=9, fmt=MOEDA)

    # Col G: receita à vista = lotes × preço × %vista
    cel(wsc, r, 7, f"=D{r}*{P('preco')}*{P('vista')}", bg=bg, h='right', size=9, fmt=MOEDA)

    # Col H: parcelas recebidas = SUMPRODUCT de meses anteriores dentro do prazo
    if m == 1:
        cel(wsc, r, 8, 0, bg=bg, h='right', size=9, fmt=MOEDA)
    else:
        # =SUMPRODUCT(($D$4:D{r-1})*($A$4:A{r-1}>=(A{r}-prazo)))*parcela
        formula_parc = (
            f"=SUMPRODUCT(($D${DATA_START}:D{r-1})"
            f"*($A${DATA_START}:A{r-1}>=(A{r}-{P('prazo')})))"
            f"*{P('parcela')}"
        )
        cel(wsc, r, 8, formula_parc, bg=bg, h='right', size=9, fmt=MOEDA)

    # Col I: total receita = G + H
    cel(wsc, r, 9, f"=G{r}+H{r}", bg=VERDE_CLA if lotes_mes(m) > 0 else bg,
        bold=True, h='right', size=9, fmt=MOEDA)

    # Col J: corretagem = lotes × preço × taxa
    cel(wsc, r, 10, f"=D{r}*{P('preco')}*{P('corret')}", bg=bg, h='right', size=9, fmt=MOEDA)

    # Col K: admin mensal = VGV × taxa_admin / 48
    cel(wsc, r, 11, f"={P('vgv')}*{P('admin')}/48", bg=bg, h='right', size=9, fmt=MOEDA)

    # Col L: contingência = (E+F+J+K) × taxa_conting
    cel(wsc, r, 12, f"=(E{r}+F{r}+J{r}+K{r})*{P('conting')}", bg=bg, h='right', size=9, fmt=MOEDA)

    # Col M: total custos = E+F+J+K+L
    cel(wsc, r, 13, f"=E{r}+F{r}+J{r}+K{r}+L{r}", bg=LARANJA, bold=True, h='right', size=9, fmt=MOEDA)

    # Col N: resultado bruto = I - M
    cel(wsc, r, 14, f"=I{r}-M{r}", bg=bg, bold=True, h='right', size=9, fmt=MOEDA)

    # Col O: pagamento terrenista = MAX(0, N) × %terr
    cel(wsc, r, 15, f"=MAX(0,N{r})*{P('terr')}", bg=bg, h='right', size=9, fmt=MOEDA)

    # Col P: resultado empreendedor = N - O
    cel(wsc, r, 16, f"=N{r}-O{r}", bg=bg, bold=True, h='right', size=9, fmt=MOEDA)

    # Col Q: impostos = I × taxa_impostos
    cel(wsc, r, 17, f"=I{r}*{P('impost')}", bg=bg, h='right', size=9, fmt=MOEDA)

    # Col R: custo financeiro = MAX(0, -acumulado_anterior) × taxa
    if m == 1:
        # Mês 1: custo financeiro sobre o capital próprio imobilizado
        cel(wsc, r, 18, f"={P('capital')}*{P('taxa_fin')}", bg=bg, h='right', size=9, fmt=MOEDA)
    else:
        cel(wsc, r, 18, f"=MAX(0,-T{r-1})*{P('taxa_fin')}", bg=bg, h='right', size=9, fmt=MOEDA)

    # Col S: fluxo líquido = P - Q - R
    fl_bg = bg  # será colorido conforme valor (Excel não faz isso automaticamente)
    cel(wsc, r, 19, f"=P{r}-Q{r}-R{r}", bg=bg, bold=True, h='right', size=9, fmt=MOEDA)

    # Col T: fluxo acumulado
    if m == 1:
        # Mês 1: S1 - capital próprio (aporte inicial)
        cel(wsc, r, 20, f"=S{r}-{P('capital')}", bg=bg, bold=True, h='right', size=9, fmt=MOEDA)
    else:
        cel(wsc, r, 20, f"=T{r-1}+S{r}", bg=bg, bold=True, h='right', size=9, fmt=MOEDA)

# Linha de totais do cronograma
r_tot = DATA_START + 48
wsc.row_dimensions[r_tot].height = 18
wsc.merge_cells(f'A{r_tot}:C{r_tot}')
cel(wsc, r_tot, 1, "TOTAL 48 MESES", bg=AZUL_ESC, bold=True, color=BRANCO, h='center', size=9)
for col in range(4, 21):  # D até T
    if col in [1,2,3]: continue
    formula = f"=SUM({get_column_letter(col)}{DATA_START}:{get_column_letter(col)}{r_tot-1})"
    if col == 20:  # acumulado: mostrar o valor final
        formula = f"=T{r_tot-1}"
    cel(wsc, r_tot, col, formula, bg=AZUL_CLA, bold=True, h='right', size=9,
        fmt=INT if col == 4 else MOEDA, thick=True)

wsc.freeze_panes = 'D4'

# ══════════════════════════════════════════════════════
# FUNÇÃO: criar aba anual referenciando o Cronograma
# ══════════════════════════════════════════════════════
COLUNAS_ANUAIS = [
    ("MÊS",               'C', None,  'left',  12),
    ("LOTES\nVEND.",      'D', INT,   'center', 9),
    ("RECEITA\nÀ VISTA",  'G', MOEDA, 'right', 14),
    ("PARCELAS\nRECEB.",  'H', MOEDA, 'right', 14),
    ("TOTAL\nRECEITA",    'I', MOEDA, 'right', 14),
    ("(-) INFRA-\nESTR.", 'E', MOEDA, 'right', 14),
    ("(-) CORRE-\nTAGEM", 'J', MOEDA, 'right', 13),
    ("(-) ADMIN\n& MKT",  None, MOEDA, 'right', 13),  # J+K+F
    ("(-) CONTIN-\nGÊNC.",'L', MOEDA, 'right', 12),
    ("TOTAL\nCUSTOS",     'M', MOEDA, 'right', 14),
    ("RESULT.\nBRUTO",    'N', MOEDA, 'right', 14),
    ("(-) TERR-\nENISTA", 'O', MOEDA, 'right', 13),
    ("(-) IMPOS-\nTOS",   'Q', MOEDA, 'right', 12),
    ("(-) CUSTO\nFINANC.",'R', MOEDA, 'right', 13),
    ("FLUXO\nLÍQUIDO",   'S', MOEDA, 'right', 14),
    ("FLUXO\nACUMUL.",    'T', MOEDA, 'right', 15),
]

def criar_aba_ano(ano_num):
    ws = wb.create_sheet(f"Ano {ano_num}")
    mes_ini = (ano_num - 1) * 12 + 1
    mes_fim = ano_num * 12
    cron_ini = DATA_START + mes_ini - 1   # linha no Cronograma
    cron_fim = DATA_START + mes_fim - 1

    # Larguras
    for i, (_, _, _, _, larg) in enumerate(COLUNAS_ANUAIS):
        ws.column_dimensions[get_column_letter(i+1)].width = larg

    # Título
    ws.merge_cells(f'A1:{get_column_letter(len(COLUNAS_ANUAIS))}1')
    c = ws['A1']
    c.value = (f"FLUXO DE CAIXA — ANO {ano_num} — "
               f"{MESES[(mes_ini-1)%12]} a {MESES[(mes_fim-1)%12]} — LOTEAMENTO PAPAGAIOS/MG")
    c.fill = fill(AZUL_ESC); c.font = Font(bold=True, size=11, color=BRANCO, name='Calibri')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24

    # Subtítulo
    ws.merge_cells(f'A2:{get_column_letter(len(COLUNAS_ANUAIS))}2')
    c = ws['A2']
    c.value = (f'=TEXT(SUMIF(Cronograma!$B${DATA_START}:Cronograma!$B${DATA_START+47},'
               f'{ano_num},Cronograma!$D${DATA_START}:Cronograma!$D${DATA_START+47}),"#,##0")'
               f'&" lotes vendidos  |  Receita: "&TEXT(SUMIF(Cronograma!$B${DATA_START}:Cronograma!$B${DATA_START+47},'
               f'{ano_num},Cronograma!$I${DATA_START}:Cronograma!$I${DATA_START+47}),"R$ #,##0")'
               f'&"  |  Fluxo Líq.: "&TEXT(SUMIF(Cronograma!$B${DATA_START}:Cronograma!$B${DATA_START+47},'
               f'{ano_num},Cronograma!$S${DATA_START}:Cronograma!$S${DATA_START+47}),"R$ #,##0")')
    c.fill = fill(AZUL_MED); c.font = Font(size=9, color=BRANCO, italic=True, name='Calibri')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 16

    # Grupos de colunas
    ws.merge_cells(f'C2:{get_column_letter(5)}2')
    ws.merge_cells(f'F2:{get_column_letter(10)}2')

    # Cabeçalho
    ws.row_dimensions[3].height = 28
    grupos_bgs = [AZUL_ESC, AZUL_ESC, "117A65","117A65","117A65",
                  "C0392B","C0392B","C0392B","C0392B","C0392B",
                  AZUL_ESC, AZUL_ESC, "1A5276","1A5276","1A5276","1A5276"]
    for i, (nome, _, _, alinha, _) in enumerate(COLUNAS_ANUAIS):
        bg = grupos_bgs[i] if i < len(grupos_bgs) else AZUL_ESC
        c = ws.cell(row=3, column=i+1, value=nome)
        c.fill = fill(bg); c.font = Font(bold=True, size=8, color=BRANCO, name='Calibri')
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = bord()

    # Dados — referências diretas ao Cronograma
    for i_mes in range(12):
        mes_abs = mes_ini + i_mes
        cron_row = DATA_START + mes_abs - 1
        r = 4 + i_mes
        ws.row_dimensions[r].height = 15
        bg = CINZA if i_mes % 2 == 0 else BRANCO

        for col_i, (_, cron_col, fmt, halign, _) in enumerate(COLUNAS_ANUAIS):
            col = col_i + 1
            if cron_col is None:
                # Admin + Marketing + F: J+K+F do Cronograma
                formula = f"=Cronograma!F{cron_row}+Cronograma!K{cron_row}"
            else:
                formula = f"=Cronograma!{cron_col}{cron_row}"

            # Cor especial para fluxo líquido e acumulado
            if cron_col == 'S':
                c = ws.cell(row=r, column=col, value=formula)
                c.fill = fill(bg); c.font = Font(bold=True, size=9, name='Calibri')
                c.alignment = Alignment(horizontal=halign, vertical='center')
                c.border = bord()
                if fmt: c.number_format = fmt
            elif cron_col == 'T':
                c = ws.cell(row=r, column=col, value=formula)
                c.fill = fill(bg); c.font = Font(bold=True, size=9, name='Calibri')
                c.alignment = Alignment(horizontal=halign, vertical='center')
                c.border = bord()
                if fmt: c.number_format = fmt
            else:
                c = ws.cell(row=r, column=col, value=formula)
                c.fill = fill(AZUL_CLA if cron_col == 'I' else bg)
                c.font = Font(bold=cron_col in ['I','M','N','S','T'], size=9, name='Calibri')
                c.alignment = Alignment(horizontal=halign, vertical='center')
                c.border = bord()
                if fmt: c.number_format = fmt

    # Total do ano
    r_tot_ano = 4 + 12
    ws.row_dimensions[r_tot_ano].height = 18
    for col_i, (_, cron_col, fmt, halign, _) in enumerate(COLUNAS_ANUAIS):
        col = col_i + 1
        if col_i == 0:
            c = ws.cell(row=r_tot_ano, column=col, value=f"TOTAL ANO {ano_num}")
            c.fill = fill(AZUL_ESC); c.font = Font(bold=True, size=9, color=BRANCO, name='Calibri')
            c.alignment = Alignment(horizontal='left', vertical='center'); c.border = bord_med()
        elif cron_col == 'T':
            # Acumulado: pegar valor do último mês
            formula = f"=Cronograma!T{cron_fim}"
            c = ws.cell(row=r_tot_ano, column=col, value=formula)
            c.fill = fill(AZUL_CLA); c.font = Font(bold=True, size=9, color=AZUL_ESC, name='Calibri')
            c.alignment = Alignment(horizontal=halign, vertical='center'); c.border = bord_med()
            if fmt: c.number_format = fmt
        elif cron_col is None:
            formula = f"=SUM({get_column_letter(col)}4:{get_column_letter(col)}{r_tot_ano-1})"
            c = ws.cell(row=r_tot_ano, column=col, value=formula)
            c.fill = fill(AZUL_CLA); c.font = Font(bold=True, size=9, color=AZUL_ESC, name='Calibri')
            c.alignment = Alignment(horizontal=halign, vertical='center'); c.border = bord_med()
            if fmt: c.number_format = fmt
        else:
            formula = f"=SUM({get_column_letter(col)}4:{get_column_letter(col)}{r_tot_ano-1})"
            c = ws.cell(row=r_tot_ano, column=col, value=formula)
            c.fill = fill(AZUL_CLA); c.font = Font(bold=True, size=9, color=AZUL_ESC, name='Calibri')
            c.alignment = Alignment(horizontal=halign, vertical='center'); c.border = bord_med()
            if fmt: c.number_format = fmt

    # Legenda
    r_leg = r_tot_ano + 2
    ws.merge_cells(f'A{r_leg}:{get_column_letter(len(COLUNAS_ANUAIS))}{r_leg}')
    c = ws[f'A{r_leg}']
    c.value = ("📌  Todos os valores são fórmulas vinculadas à aba Cronograma, "
               "que por sua vez vincula com Parâmetros. "
               "Para alterar qualquer premissa, edite apenas a aba Parâmetros.")
    c.fill = fill(AMARELO); c.font = Font(size=8, italic=True, color="7D6608", name='Calibri')
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    c.border = bord(); ws.row_dimensions[r_leg].height = 22

    ws.freeze_panes = 'B4'

for ano in range(1, 5):
    criar_aba_ano(ano)

# ══════════════════════════════════════════════════════
# ABA RESUMO ANUAL (referências às abas anuais)
# ══════════════════════════════════════════════════════
wsr = wb.create_sheet("Resumo Anual", 0)
wsr.column_dimensions['A'].width = 38
for l in ['B','C','D','E','F']:
    wsr.column_dimensions[l].width = 20

wsr.merge_cells('A1:F1')
c = wsr['A1']; c.value = "RESUMO ANUAL — FLUXO DE CAIXA — LOTEAMENTO PAPAGAIOS/MG"
c.fill = fill(AZUL_ESC); c.font = Font(bold=True, size=13, color=BRANCO, name='Calibri')
c.alignment = Alignment(horizontal='center', vertical='center')
wsr.row_dimensions[1].height = 30

# Cabeçalhos
for i, (h, bg) in enumerate([("INDICADOR",VERDE_ESC),("ANO 1",AZUL_ESC),
                               ("ANO 2",AZUL_ESC),("ANO 3",AZUL_ESC),
                               ("ANO 4",AZUL_ESC),("TOTAL 4 ANOS",VERDE_ESC)]):
    c = wsr.cell(row=2, column=i+1, value=h)
    c.fill = fill(bg); c.font = Font(bold=True, size=10, color=BRANCO, name='Calibri')
    c.alignment = Alignment(horizontal='center' if i > 0 else 'left', vertical='center')
    c.border = bord()
wsr.row_dimensions[2].height = 18

# Linha de SUMIF por ano para cada indicador (referencia o Cronograma diretamente)
def sumif_ano(ano, col_cron):
    return (f"=SUMIF(Cronograma!$B${DATA_START}:Cronograma!$B${DATA_START+47},"
            f"{ano},Cronograma!${col_cron}${DATA_START}:Cronograma!${col_cron}${DATA_START+47})")

def acum_final_ano(ano):
    r = DATA_START + ano*12 - 1
    return f"=Cronograma!$T${r}"

linhas_res = [
    ("Lotes vendidos",              'D', INT,   False, False),
    ("Receita à vista",             'G', MOEDA, False, False),
    ("Parcelas recebidas",          'H', MOEDA, False, False),
    ("TOTAL RECEITA",               'I', MOEDA, True,  False),
    ("(-) Infraestrutura",          'E', MOEDA, False, False),
    ("(-) Corretagem",              'J', MOEDA, False, False),
    ("(-) Admin & Marketing",       None, MOEDA, False, False),  # F+K
    ("(-) Contingência",            'L', MOEDA, False, False),
    ("TOTAL CUSTOS",                'M', MOEDA, True,  False),
    ("Resultado Bruto",             'N', MOEDA, True,  False),
    ("(-) Pagto. Terrenista",       'O', MOEDA, False, False),
    ("(-) Impostos",                'Q', MOEDA, False, False),
    ("(-) Custo Financeiro",        'R', MOEDA, False, False),
    ("FLUXO LÍQUIDO EMPREENDEDOR",  'S', MOEDA, True,  False),
    ("SALDO ACUMULADO (final ano)", 'T', MOEDA, True,  True),  # usa acum_final
]

row = 3
for nome, col_c, fmt, bold, use_acum in linhas_res:
    wsr.row_dimensions[row].height = 16
    bg_n = CINZA if row%2==0 else BRANCO
    cel(wsr, row, 1, nome, bg=bg_n, bold=bold, h='left', size=9)

    total_formula_parts = []
    for i_ano, ano in enumerate([1,2,3,4]):
        col = i_ano + 2
        if use_acum:
            f = acum_final_ano(ano)
        elif col_c is None:
            f = (f"=SUMIF(Cronograma!$B${DATA_START}:Cronograma!$B${DATA_START+47},{ano},"
                 f"Cronograma!$F${DATA_START}:Cronograma!$F${DATA_START+47})"
                 f"+SUMIF(Cronograma!$B${DATA_START}:Cronograma!$B${DATA_START+47},{ano},"
                 f"Cronograma!$K${DATA_START}:Cronograma!$K${DATA_START+47})")
        else:
            f = sumif_ano(ano, col_c)
        c = wsr.cell(row=row, column=col, value=f)
        bg_v = VERDE_CLA if bold else bg_n
        c.fill = fill(bg_v); c.font = Font(bold=bold, size=9, name='Calibri')
        c.alignment = Alignment(horizontal='right', vertical='center')
        c.border = bord()
        if fmt: c.number_format = fmt
        if not use_acum:
            total_formula_parts.append(f'{get_column_letter(col)}{row}')

    # Coluna F: total 4 anos
    col_tot = 6
    if use_acum:
        f_tot = acum_final_ano(4)
    else:
        f_tot = "="+"+".join(total_formula_parts)
    c = wsr.cell(row=row, column=col_tot, value=f_tot)
    c.fill = fill(VERDE_CLA if bold else bg_n)
    c.font = Font(bold=True, size=9, color=VERDE_ESC if bold else PRETO, name='Calibri')
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.border = bord()
    if fmt: c.number_format = fmt
    row += 1

# Indicadores chave (referência a Parâmetros)
row += 1
wsr.merge_cells(f'A{row}:F{row}')
c = wsr[f'A{row}']; c.value = "INDICADORES CHAVE (vínculos com Parâmetros)"
c.fill = fill(AZUL_ESC); c.font = Font(bold=True, size=10, color=BRANCO, name='Calibri')
c.alignment = Alignment(horizontal='center', vertical='center')
c.border = bord(); wsr.row_dimensions[row].height = 20; row += 1

indicadores_res = [
    ("VGV Total",                    f"={P('vgv')}",        MOEDA),
    ("Preço por lote",               f"={P('preco')}",      MOEDA),
    ("Total de lotes",               f"={P('lotes')}",      INT),
    ("% Empreendedor",               f"={P('emp')}",        PERC),
    ("% Terrenista",                 f"={P('terr')}",       PERC),
    ("Capital próprio investido",    f"={P('capital')}",    MOEDA),
    ("Infra total",                  f"={P('infra')}",      MOEDA),
]
for nome_i, val_i, fmt_i in indicadores_res:
    wsr.row_dimensions[row].height = 16
    cel(wsr, row, 1, nome_i, bg=CINZA if row%2==0 else BRANCO, h='left', size=9)
    wsr.merge_cells(f'B{row}:F{row}')
    c = wsr.cell(row=row, column=2, value=val_i)
    c.fill = fill(AZUL_CLA); c.font = Font(bold=True, size=10, color=AZUL_ESC, name='Calibri')
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.border = bord()
    if fmt_i: c.number_format = fmt_i
    row += 1

wsr.freeze_panes = 'B3'

# ══════════════════════════════════════════════════════
# SALVAR
# ══════════════════════════════════════════════════════
path = "/Users/leonardoforx/Downloads/Papagaios/FluxoCaixa_Mensal_Papagaios.xlsx"
wb.save(path)
print(f"✅ Fluxo de caixa salvo: {path}")
print(f"\n📐 ARQUITETURA:")
print(f"   Parâmetros  → edite aqui (preço, taxas, %)")
print(f"   Cronograma  → 48 meses com fórmulas ligadas a Parâmetros")
print(f"   Ano 1-4     → referências diretas ao Cronograma")
print(f"   Resumo Anual → SUMIF por ano sobre o Cronograma")
print(f"\n   Altere qualquer célula amarela nos Parâmetros e tudo recalcula.")
