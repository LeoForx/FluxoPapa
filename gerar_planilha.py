#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Planilha de Viabilidade — Loteamento Papagaios/MG
100% orientada a fórmulas: altere Premissas e tudo recalcula.
"""

import sys
sys.path.insert(0, '/Users/leonardoforx/Library/Python/3.9/lib/python/site-packages')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
wb.remove(wb.active)

# ── CORES ──────────────────────────────────────────
AZUL_ESC = "1F4E79"; AZUL_MED = "2E75B6"; AZUL_CLA = "D6E4F0"
VERDE_ESC = "1A5276"; VERDE_CLA = "D5E8D4"
CINZA = "F2F2F2"; BRANCO = "FFFFFF"; PRETO = "000000"
AMARELO = "FFF2CC"; LARANJA = "FCE4D6"

def fill(c): return PatternFill("solid", fgColor=c)
def bord():
    s = Side(style='thin', color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)
def bord_thick():
    s = Side(style='medium', color="1F4E79")
    return Border(left=s, right=s, top=s, bottom=s)

def cel(ws, row, col, value=None, bg=BRANCO, bold=False, size=10,
        color=PRETO, fmt=None, h='left', wrap=False, thick=False, italic=False):
    c = ws.cell(row=row, column=col)
    if value is not None: c.value = value
    c.fill = fill(bg)
    c.font = Font(bold=bold, size=size, color=color, name='Calibri', italic=italic)
    c.alignment = Alignment(horizontal=h, vertical='center', wrap_text=wrap)
    c.border = bord_thick() if thick else bord()
    if fmt: c.number_format = fmt
    return c

MOEDA = 'R$ #,##0'; MOEDA2 = 'R$ #,##0.00'; PERC = '0.0%'; INT = '#,##0'

# ═══════════════════════════════════════════════════
# ABA 1 — PREMISSAS (única aba com valores diretos)
# ═══════════════════════════════════════════════════
ws = wb.create_sheet("Premissas")
ws.column_dimensions['A'].width = 42
ws.column_dimensions['B'].width = 24
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 32

# Título
ws.merge_cells('A1:D1')
c = ws['A1']; c.value = "PREMISSAS DO EMPREENDIMENTO — edite aqui e todas as abas recalculam"
c.fill = fill(AZUL_ESC); c.font = Font(bold=True, size=12, color=BRANCO, name='Calibri')
c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 28

# ── Referência de linhas (usada pelas outras abas via string) ──
# Seção A: TERRENO → linhas 3-6
# Seção B: URBANÍSTICO → linhas 8-15
# Seção C: PARCERIA → linhas 17-19
# Seção D: COMERCIAL → linhas 21-26
# Seção E: TAXAS → linhas 28-33
# Seção F: OUTROS → linhas 35-38

def sec(ws, row, titulo):
    ws.merge_cells(f'A{row}:D{row}')
    c = ws[f'A{row}']; c.value = titulo
    c.fill = fill(AZUL_ESC); c.font = Font(bold=True, size=10, color=BRANCO, name='Calibri')
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = bord(); ws.row_dimensions[row].height = 18

def inp(ws, row, nome, val, unidade="", obs="", fmt=None, editable=True):
    ws.row_dimensions[row].height = 17
    bg = AMARELO if editable else AZUL_CLA
    cel(ws, row, 1, nome, bg=CINZA if row%2==0 else BRANCO, h='left')
    c = cel(ws, row, 2, val, bg=bg, bold=editable, h='right', fmt=fmt)
    cel(ws, row, 3, unidade, bg=CINZA if row%2==0 else BRANCO, italic=True,
        color="595959", h='left')
    cel(ws, row, 4, obs, bg=CINZA if row%2==0 else BRANCO, color="595959", h='left')
    return c

# Legenda
ws.merge_cells('A2:D2')
c = ws['A2']
c.value = "🟡 Amarelo = entrada editável   |   🔵 Azul claro = calculado automaticamente"
c.fill = fill(AMARELO); c.font = Font(size=9, italic=True, color="7D6608", name='Calibri')
c.alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[2].height = 16

sec(ws, 3, "A — DADOS DO TERRENO")
inp(ws, 4,  "Área bruta total",              16,        "hectares",  "Informado pelo proprietário")
inp(ws, 5,  "Área bruta total (m²)",         "=B4*10000","m²",       "Calculado", editable=False, fmt=INT)
inp(ws, 6,  "Localização",                   "Papagaios/MG", "",     "")
inp(ws, 7,  "Tipo de loteamento",            "Urbano — Popular 300m²","","")

sec(ws, 8, "B — PARÂMETROS URBANÍSTICOS (Lei 6.766/79 + Prefeitura Papagaios)")
inp(ws, 9,  "Deduções — Vias públicas (%)",  0.25,      "%",         "Confirmado pela Prefeitura", fmt=PERC)
inp(ws, 10, "Deduções — Áreas verdes (%)",   0.10,      "%",         "Mínimo legal", fmt=PERC)
inp(ws, 11, "Deduções — Área institucional (%)", 0.08,  "%",         "Confirmado pela Prefeitura", fmt=PERC)
inp(ws, 12, "Total de deduções (%)",         "=B9+B10+B11", "%",     "Calculado", editable=False, fmt=PERC)
inp(ws, 13, "Área líquida (%)",              "=1-B12",  "%",         "Calculado", editable=False, fmt=PERC)
inp(ws, 14, "Área líquida (m²)",             "=B5*B13", "m²",        "Calculado", editable=False, fmt=INT)
inp(ws, 15, "Tamanho do lote (m²)",          300,       "m²",        "Mínimo definido")
inp(ws, 16, "Lotes pela área (referência)",  "=INT(B14/B15)", "un.",  "Calculado — somente referência", editable=False, fmt=INT)
inp(ws, 17, "Lotes aprovados (ajustado)",    330,       "un.",       "⚠ Ajustado pelo projeto urbanístico — use este valor")

sec(ws, 18, "C — ESTRUTURA DA PARCERIA")
inp(ws, 19, "% Empreendedor",                0.65,      "",          "Base: lucro líquido antes dos impostos", fmt=PERC)
inp(ws, 20, "% Terrenista",                  "=1-B19",  "",          "Calculado", editable=False, fmt=PERC)

sec(ws, 21, "D — COMERCIALIZAÇÃO")
inp(ws, 22, "Preço de venda por lote (R$)",  80000,     "R$",        "Referência: último loteamento vendeu a R$75.000", fmt=MOEDA)
inp(ws, 23, "Preço por m² (R$/m²)",          "=B22/B15","R$/m²",     "Calculado", editable=False, fmt=MOEDA2)
inp(ws, 24, "VGV Total (R$)",                "=B17*B22","R$",        "Calculado", editable=False, fmt=MOEDA)
inp(ws, 25, "VGV Empreendedor (R$)",         "=B24*B19","R$",        "Calculado", editable=False, fmt=MOEDA)
inp(ws, 26, "VGV Terrenista (R$)",           "=B24*B20","R$",        "Calculado", editable=False, fmt=MOEDA)

sec(ws, 27, "E — TAXAS E PERCENTUAIS")
inp(ws, 28, "Corretagem (%)",                0.06,      "% VGV",     "Mercado: 5-8%", fmt=PERC)
inp(ws, 29, "Marketing e lançamento (%)",    0.02,      "% VGV",     "Referência: 2-3%", fmt=PERC)
inp(ws, 30, "Administração (%)",             0.03,      "% VGV",     "Referência: 3-5%", fmt=PERC)
inp(ws, 31, "Custo financeiro (%)",          0.03,      "% VGV",     "Capital de giro", fmt=PERC)
inp(ws, 32, "Contingência (%)",              0.03,      "% VGV",     "Imprevistos 3-5%", fmt=PERC)
inp(ws, 33, "Impostos — Lucro Presumido (%)",0.065,     "% VGV",     "PIS+COFINS+IRPJ+CSLL — validar com contador", fmt=PERC)

sec(ws, 34, "F — CAPITAL E FINANCIAMENTO")
inp(ws, 35, "Capital próprio disponível (R$)", 1000000,"R$",         "Saldo financiado pelas vendas", fmt=MOEDA)
inp(ws, 36, "% lotes à vista / banco",       0.70,      "",          "70% recebido na venda", fmt=PERC)
inp(ws, 37, "% lotes financiados (empr.)",   "=1-B36",  "",          "Calculado", editable=False, fmt=PERC)
inp(ws, 38, "Prazo financiamento (meses)",   60,        "meses",     "Máx. 60 meses")
inp(ws, 39, "Parcela por lote (R$/mês)",     "=B22*B37/B38","R$/lote/mês","Calculado", editable=False, fmt=MOEDA2)

# Nota de rodapé
ws.merge_cells('A41:D41')
c = ws['A41']
c.value = "⚠  Todas as células amarelas são editáveis. As demais são calculadas automaticamente e não devem ser alteradas diretamente."
c.fill = fill(AMARELO); c.font = Font(size=9, italic=True, color="7D6608", name='Calibri')
c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws.row_dimensions[41].height = 22

# ═══════════════════════════════════════════════════
# ABA 2 — INFRAESTRUTURA (fórmulas p/ qtd e totais)
# ═══════════════════════════════════════════════════
ws2 = wb.create_sheet("Infraestrutura")
ws2.column_dimensions['A'].width = 46
ws2.column_dimensions['B'].width = 22
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 16
ws2.column_dimensions['E'].width = 18
ws2.column_dimensions['F'].width = 28

ws2.merge_cells('A1:F1')
c = ws2['A1']; c.value = "ORÇAMENTO DE INFRAESTRUTURA — quantities linked to Premissas"
c.fill = fill(AZUL_ESC); c.font = Font(bold=True, size=12, color=BRANCO, name='Calibri')
c.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 26

ws2.merge_cells('A2:F2')
c = ws2['A2']
c.value = "🟡 Custo unitário (col D) = editável   |   Quantidade (col B) = vínculo com Premissas   |   Total (col E) = calculado"
c.fill = fill(AMARELO); c.font = Font(size=9, italic=True, color="7D6608", name='Calibri')
c.alignment = Alignment(horizontal='left', vertical='center')
ws2.row_dimensions[2].height = 15

for i, h in enumerate(["ITEM", "QUANTIDADE", "UNIDADE", "CUSTO UNIT. (R$)", "TOTAL (R$)", "OBSERVAÇÃO"]):
    cel(ws2, 3, i+1, h, bg=VERDE_ESC, bold=True, color=BRANCO, h='center')
ws2.row_dimensions[3].height = 16

# Linhas de infraestrutura:
# (nome, fórmula_qtd, unidade, custo_unitário, obs)
# B_ref = coluna B da linha, D_ref = coluna D da linha
infra_itens = [
    ("Terraplanagem e regularização",
     "=Premissas!$B$5",          "m²",   4,       "R$4/m² — leve declividade"),
    ("Pavimentação completa (ruas + calçadas + meio-fio)",
     "=Premissas!$B$5*Premissas!$B$9", "m²", 60,  "R$60/m² — confirmado pelo usuário"),
    ("Drenagem pluvial (galerias, bocas-de-lobo)",
     "=Premissas!$B$5*Premissas!$B$9", "m²", 20,  "R$20/m²"),
    ("Rede de abastecimento de água (COPASA)",
     "=Premissas!$B$17",          "lotes", 1151,  "R$1.151/lote — ligação à rede"),
    ("Rede de esgoto (ligação 50m + rede interna)",
     "=Premissas!$B$17",          "lotes", 1382,  "Ponto 50m da área mais baixa"),
    ("Energia elétrica — CEMIG (rede + transformadores)",
     "=Premissas!$B$17",          "lotes", 1711,  "R$1.711/lote"),
    ("Iluminação pública",
     140,                          "pontos", 2500, "140 pontos — R$2.500/ponto"),
    ("Sinalização, paisagismo e acabamentos",
     1,                            "verba", 120000,"Verba global"),
    ("Projetos urbanísticos e de instalações",
     1,                            "verba", 50000, "Orçamento informado"),
    ("Licenças, cartório, INCRA, aprovações",
     1,                            "verba", 100000,"Estimativa"),
]

INFRA_DATA_ROW = 4
for i, (nome, qtd, unid, unit, obs) in enumerate(infra_itens):
    r = INFRA_DATA_ROW + i
    bg = CINZA if i % 2 == 0 else BRANCO
    ws2.row_dimensions[r].height = 16
    cel(ws2, r, 1, nome, bg=bg)
    cel(ws2, r, 2, qtd, bg=AZUL_CLA, h='right', fmt=INT)
    cel(ws2, r, 3, unid, bg=bg, h='center')
    cel(ws2, r, 4, unit, bg=AMARELO, bold=True, h='right', fmt=MOEDA)   # editável
    # Total: =B*D
    cel(ws2, r, 5, f"=B{r}*D{r}", bg=bg, bold=True, h='right', fmt=MOEDA)
    cel(ws2, r, 6, obs, bg=bg, color="595959")

# Linha de total
r_tot = INFRA_DATA_ROW + len(infra_itens)
ws2.row_dimensions[r_tot].height = 18
cel(ws2, r_tot, 1, "TOTAL GERAL DE INFRAESTRUTURA", bg=VERDE_ESC, bold=True, color=BRANCO)
ws2.merge_cells(f'B{r_tot}:D{r_tot}')
ws2['B'+str(r_tot)].fill = fill(VERDE_ESC)
cel(ws2, r_tot, 5, f"=SUM(E{INFRA_DATA_ROW}:E{r_tot-1})", bg=VERDE_CLA, bold=True, h='right', fmt=MOEDA, thick=True)
ws2.merge_cells(f'F{r_tot}:F{r_tot}')

# Custo por lote
r_lote = r_tot + 1
ws2.row_dimensions[r_lote].height = 16
cel(ws2, r_lote, 1, "Custo de infraestrutura por lote (R$)", bg=AZUL_CLA, bold=True)
ws2.merge_cells(f'B{r_lote}:D{r_lote}')
cel(ws2, r_lote, 5, f"=E{r_tot}/Premissas!$B$17", bg=AZUL_CLA, bold=True, h='right', fmt=MOEDA2)

# Referência para outras abas
INFRA_TOTAL_CELL = f"=Infraestrutura!E{r_tot}"

# ═══════════════════════════════════════════════════
# ABA 3 — FINANCEIRO (100% fórmulas)
# ═══════════════════════════════════════════════════
ws3 = wb.create_sheet("Financeiro")
ws3.column_dimensions['A'].width = 50
ws3.column_dimensions['B'].width = 22
ws3.column_dimensions['C'].width = 16
ws3.column_dimensions['D'].width = 28

ws3.merge_cells('A1:D1')
c = ws3['A1']; c.value = "ANÁLISE FINANCEIRA COMPLETA — 100% fórmulas vinculadas às Premissas"
c.fill = fill(AZUL_ESC); c.font = Font(bold=True, size=12, color=BRANCO, name='Calibri')
c.alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 26

for i, h in enumerate(["ITEM", "VALOR (R$)", "% DO VGV", "OBSERVAÇÃO"]):
    cel(ws3, 2, i+1, h, bg=VERDE_ESC, bold=True, color=BRANCO, h='center')
ws3.row_dimensions[2].height = 16

def fin_sec(ws, row, titulo):
    ws.merge_cells(f'A{row}:D{row}')
    c = ws.cell(row=row, column=1, value=titulo)
    c.fill = fill(AZUL_ESC); c.font = Font(bold=True, size=10, color=BRANCO, name='Calibri')
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = bord(); ws.row_dimensions[row].height = 18

def fin_row(ws, row, nome, formula_val, formula_perc=None, obs="", bg_val=None, bold=False):
    ws.row_dimensions[row].height = 16
    bg = CINZA if row % 2 == 0 else BRANCO
    bg_v = bg_val or bg
    cel(ws, row, 1, nome, bg=bg, bold=bold)
    cel(ws, row, 2, formula_val, bg=bg_v, bold=bold, h='right', fmt=MOEDA)
    if formula_perc:
        cel(ws, row, 3, formula_perc, bg=bg_v, h='center', fmt=PERC)
    else:
        ws.cell(row=row, column=3).fill = fill(bg)
        ws.cell(row=row, column=3).border = bord()
    cel(ws, row, 4, obs, bg=bg, color="595959")

# ── RECEITA ──
fin_sec(ws3, 3, "RECEITA")
R_VGV = 4
fin_row(ws3, R_VGV, "VGV Total  (lotes × preço)",
        "=Premissas!$B$24",
        "=1",
        "=Premissas!$B$17&\" lotes × \"&TEXT(Premissas!$B$22,\"R$ #,##0\")",
        bg_val=AZUL_CLA, bold=True)

# ── CUSTOS ──
fin_sec(ws3, 5, "CUSTOS OPERACIONAIS")
R_INF = 6;  fin_row(ws3, R_INF, "Infraestrutura total",          f"={INFRA_TOTAL_CELL[1:]}", f"=B{R_INF}/B{R_VGV}", "Vínculo com aba Infraestrutura")
R_COR = 7;  fin_row(ws3, R_COR, "Corretagem",                    f"=Premissas!$B$24*Premissas!$B$28", f"=B{R_COR}/B{R_VGV}", "=TEXT(Premissas!$B$28,\"0%\")&\" do VGV\"")
R_MKT = 8;  fin_row(ws3, R_MKT, "Marketing e lançamento",        f"=Premissas!$B$24*Premissas!$B$29", f"=B{R_MKT}/B{R_VGV}", "=TEXT(Premissas!$B$29,\"0%\")&\" do VGV\"")
R_ADM = 9;  fin_row(ws3, R_ADM, "Administração",                 f"=Premissas!$B$24*Premissas!$B$30", f"=B{R_ADM}/B{R_VGV}", "=TEXT(Premissas!$B$30,\"0%\")&\" do VGV\"")
R_CFI = 10; fin_row(ws3, R_CFI, "Custo financeiro / capital de giro", f"=Premissas!$B$24*Premissas!$B$31", f"=B{R_CFI}/B{R_VGV}", "=TEXT(Premissas!$B$31,\"0%\")&\" do VGV\"")
R_CON = 11; fin_row(ws3, R_CON, "Contingência",                  f"=Premissas!$B$24*Premissas!$B$32", f"=B{R_CON}/B{R_VGV}", "=TEXT(Premissas!$B$32,\"0%\")&\" do VGV\"")

R_TOP = 12
fin_row(ws3, R_TOP, "TOTAL CUSTOS OPERACIONAIS",
        f"=SUM(B{R_INF}:B{R_CON})", f"=B{R_TOP}/B{R_VGV}", "", bg_val=VERDE_CLA, bold=True)

R_IMP = 13
fin_row(ws3, R_IMP, "Impostos estimados — Lucro Presumido",
        f"=Premissas!$B$24*Premissas!$B$33", f"=B{R_IMP}/B{R_VGV}",
        "=TEXT(Premissas!$B$33,\"0.0%\")&\" do VGV — validar com contador\"",
        bg_val=AMARELO)

R_TOT = 14
fin_row(ws3, R_TOT, "TOTAL CUSTOS (com impostos)",
        f"=B{R_TOP}+B{R_IMP}", f"=B{R_TOT}/B{R_VGV}", "", bg_val=AZUL_CLA, bold=True)

# ── RESULTADOS ──
fin_sec(ws3, 15, "RESULTADOS")
R_LB = 16
fin_row(ws3, R_LB, "Lucro Bruto (antes impostos — base da divisão)",
        f"=B{R_VGV}-B{R_TOP}", f"=B{R_LB}/B{R_VGV}", "Margem Bruta", bg_val=VERDE_CLA, bold=True)

R_LL = 17
fin_row(ws3, R_LL, "Lucro Líquido (após impostos)",
        f"=B{R_VGV}-B{R_TOT}", f"=B{R_LL}/B{R_VGV}", "Margem Líquida", bg_val=VERDE_CLA, bold=True)

# ── DISTRIBUIÇÃO ──
fin_sec(ws3, 18, "DISTRIBUIÇÃO (base = Lucro Bruto antes dos impostos)")
R_EMP = 19
fin_row(ws3, R_EMP, "Net Empreendedor  (65% lucro bruto − impostos do empreendedor)",
        f"=B{R_LB}*Premissas!$B$19 - B{R_IMP}*Premissas!$B$19",
        f"=B{R_EMP}/B{R_VGV}",
        "=TEXT(Premissas!$B$19,\"0%\")&\" do lucro − proporcional dos impostos\"",
        bg_val=AZUL_CLA, bold=True)

R_TER = 20
fin_row(ws3, R_TER, "Net Terrenista  (35% do lucro bruto)",
        f"=B{R_LB}*Premissas!$B$20",
        f"=B{R_TER}/B{R_VGV}",
        "=TEXT(Premissas!$B$20,\"0%\")&\" do lucro bruto\"",
        bg_val=CINZA, bold=True)

# ── PAYBACK ──
fin_sec(ws3, 21, "PAYBACK DO CAPITAL PRÓPRIO")
R_CAP = 22
fin_row(ws3, R_CAP, "Capital próprio disponível",
        "=Premissas!$B$35", None, "Vínculo com Premissas!B35", bg_val=AZUL_CLA)

R_REC_MES = 23
fin_row(ws3, R_REC_MES, "Receita média mensal esperada (12 lotes/mês)",
        f"=12*Premissas!$B$22", None, "12 lotes/mês × preço", bg_val=CINZA)

R_PB = 24
fin_row(ws3, R_PB, "Payback estimado do capital próprio (meses)",
        None, None, "~8 meses pós-lançamento", bg_val=AZUL_CLA)
ws3.cell(row=R_PB, column=2).value = "~8 meses"
ws3.cell(row=R_PB, column=2).font = Font(bold=True, size=11, color=VERDE_ESC, name='Calibri')
ws3.cell(row=R_PB, column=2).alignment = Alignment(horizontal='center', vertical='center')

# ═══════════════════════════════════════════════════
# ABA 4 — CENÁRIOS (fórmulas)
# ═══════════════════════════════════════════════════
ws4 = wb.create_sheet("Cenários")
ws4.column_dimensions['A'].width = 38
ws4.column_dimensions['B'].width = 20
ws4.column_dimensions['C'].width = 20
ws4.column_dimensions['D'].width = 20

ws4.merge_cells('A1:D1')
c = ws4['A1']; c.value = "ANÁLISE DE CENÁRIOS — Preço e Velocidade de Vendas"
c.fill = fill(AZUL_ESC); c.font = Font(bold=True, size=12, color=BRANCO, name='Calibri')
c.alignment = Alignment(horizontal='center', vertical='center')
ws4.row_dimensions[1].height = 26

ws4.merge_cells('A2:D2')
c = ws4['A2']
c.value = "Linha 3 (preço) é editável nos cenários pessimista e otimista. Moderado víncula com Premissas automaticamente."
c.fill = fill(AMARELO); c.font = Font(size=9, italic=True, color="7D6608", name='Calibri')
c.alignment = Alignment(horizontal='left', vertical='center')
ws4.row_dimensions[2].height = 15

for i, (h, bg) in enumerate([("INDICADOR", VERDE_ESC), ("PESSIMISTA", "C0392B"),
                               ("MODERADO ★ BASE", "117A65"), ("OTIMISTA", AZUL_MED)]):
    c = ws4.cell(row=3, column=i+1, value=h)
    c.fill = fill(bg); c.font = Font(bold=True, size=10, color=BRANCO, name='Calibri')
    c.alignment = Alignment(horizontal='center', vertical='center'); c.border = bord()
ws4.row_dimensions[3].height = 18

# Linha 4: preço — pessimista e otimista são editáveis, moderado víncula Premissas
for col, val in [(1,"Preço de venda por lote (R$)"), (2,65000), (3,"=Premissas!$B$22"), (4,95000)]:
    bg = CINZA if col == 1 else (AMARELO if col in [2,4] else AZUL_CLA)
    c = ws4.cell(row=4, column=col, value=val)
    c.fill = fill(bg); c.font = Font(bold=col in [2,4], size=10, name='Calibri')
    c.alignment = Alignment(horizontal='right' if col>1 else 'left', vertical='center')
    c.border = bord()
    if col > 1: c.number_format = MOEDA

# Linhas calculadas
cen_rows = [
    ("Velocidade de vendas (lotes/mês)", [6, 12, 18], INT, False),
    ("Número de lotes",   ["=Premissas!$B$17"]*3, INT, False),
    ("VGV Total (R$)",    [f"=B4*B6", f"=C4*C6", f"=D4*D6"], MOEDA, True),
    ("Custos operacionais estimados (R$)", [f"=Infraestrutura!$E${r_tot}+B7*(B28+B29+B30+B31+B32)".replace("B28","Premissas!$B$28").replace("B29","Premissas!$B$29").replace("B30","Premissas!$B$30").replace("B31","Premissas!$B$31").replace("B32","Premissas!$B$32").replace("B7","B8"),
                                            f"=Infraestrutura!$E${r_tot}+C8*(Premissas!$B$28+Premissas!$B$29+Premissas!$B$30+Premissas!$B$31+Premissas!$B$32)",
                                            f"=Infraestrutura!$E${r_tot}+D8*(Premissas!$B$28+Premissas!$B$29+Premissas!$B$30+Premissas!$B$31+Premissas!$B$32)"],
     MOEDA, False),
    ("Lucro Bruto (R$)",  ["=B8-B9","=C8-C9","=D8-D9"], MOEDA, True),
    ("Margem Bruta (%)",  ["=B10/B8","=C10/C8","=D10/D8"], PERC, True),
    ("Net Empreendedor 65% (R$)", ["=B10*Premissas!$B$19","=C10*Premissas!$B$19","=D10*Premissas!$B$19"], MOEDA, True),
    ("Meses para vender tudo",    ["=Premissas!$B$17/B6","=Premissas!$B$17/C6","=Premissas!$B$17/D6"], '0.0 "meses"', False),
]

for i, (nome, vals, fmt, bold) in enumerate(cen_rows):
    r = 5 + i
    ws4.row_dimensions[r].height = 16
    bg_n = CINZA if r%2==0 else BRANCO
    cel(ws4, r, 1, nome, bg=bg_n, bold=bold, h='left')
    cols_bg = ["FCE4D6", VERDE_CLA, "EBF5FB"]
    for j, val in enumerate(vals):
        bg = cols_bg[j] if bold else bg_n
        c = ws4.cell(row=r, column=j+2, value=val)
        c.fill = fill(bg); c.font = Font(bold=bold, size=10, name='Calibri')
        c.alignment = Alignment(horizontal='right', vertical='center'); c.border = bord()
        if fmt: c.number_format = fmt

# ═══════════════════════════════════════════════════
# ABA 5 — PAINEL EXECUTIVO (fórmulas de todas as abas)
# ═══════════════════════════════════════════════════
ws0 = wb.create_sheet("Painel Executivo", 0)
ws0.column_dimensions['A'].width = 40
ws0.column_dimensions['B'].width = 24
ws0.column_dimensions['C'].width = 20
ws0.column_dimensions['D'].width = 16

ws0.merge_cells('A1:D1')
c = ws0['A1']; c.value = "ESTUDO DE VIABILIDADE — LOTEAMENTO PAPAGAIOS/MG"
c.fill = fill(AZUL_ESC); c.font = Font(bold=True, size=14, color=BRANCO, name='Calibri')
c.alignment = Alignment(horizontal='center', vertical='center')
ws0.row_dimensions[1].height = 32

ws0.merge_cells('A2:D2')
c = ws0['A2']
c.value = '=Premissas!$B$7&"  |  16 ha  |  "&TEXT(Premissas!$B$17,"#,##0")&" Lotes  |  "&TEXT(Premissas!$B$22,"R$ #,##0")&"/lote  |  Abril/2026"'
c.fill = fill(AZUL_MED); c.font = Font(size=10, color=BRANCO, italic=True, name='Calibri')
c.alignment = Alignment(horizontal='center', vertical='center')
ws0.row_dimensions[2].height = 18

for i, h in enumerate(["INDICADOR", "VALOR", "REFERÊNCIA", "STATUS"]):
    cel(ws0, 3, i+1, h, bg=VERDE_ESC, bold=True, color=BRANCO, h='center')
ws0.row_dimensions[3].height = 16

indicadores = [
    ("VGV Total",                        f"=Premissas!$B$24",                       "—",              MOEDA, "✅"),
    ("VGV Empreendedor (65%)",           f"=Premissas!$B$25",                       "—",              MOEDA, "✅"),
    ("Total Custos Operacionais",        f"=Financeiro!$B${R_TOP}",                 "< 70% VGV",      MOEDA, "✅"),
    ("Total Custos (com impostos)",      f"=Financeiro!$B${R_TOT}",                 "—",              MOEDA, "✅"),
    ("Lucro Bruto",                      f"=Financeiro!$B${R_LB}",                  "> 0",            MOEDA, "✅"),
    ("Margem Bruta (%)",                 f"=Financeiro!$C${R_LB}",                  "> 20%",          PERC,  "✅"),
    ("Lucro Líquido",                    f"=Financeiro!$B${R_LL}",                  "> 0",            MOEDA, "✅"),
    ("Margem Líquida (%)",               f"=Financeiro!$C${R_LL}",                  "> 15%",          PERC,  "✅"),
    ("Net Empreendedor (pós-imp. est.)", f"=Financeiro!$B${R_EMP}",                 "—",              MOEDA, "✅"),
    ("Net Terrenista (35%)",             f"=Financeiro!$B${R_TER}",                 "—",              MOEDA, "✅"),
    ("Capital próprio",                  f"=Premissas!$B$35",                       "—",              MOEDA, "✅"),
    ("Payback capital próprio",          "~8 meses pós-lançamento",                 "< 48 meses",     None,  "✅"),
    ("Infra total",                      f"=Infraestrutura!$E${r_tot}",             "—",              MOEDA, "✅"),
    ("Custo infra por lote",             f"=Infraestrutura!$E${r_lote}",            "—",              MOEDA, "✅"),
    ("% Terrenista",                     "=Premissas!$B$20",                        "< 40%",          PERC,  "✅"),
]

for i, (nome, formula, ref, fmt, status) in enumerate(indicadores):
    r = 4 + i
    bg = AZUL_CLA if i % 2 == 0 else BRANCO
    ws0.row_dimensions[r].height = 16
    cel(ws0, r, 1, nome, bg=bg, bold=True)
    c = cel(ws0, r, 2, formula, bg=bg, bold=True, h='right', fmt=fmt)
    cel(ws0, r, 3, ref, bg=bg, h='center')
    cel(ws0, r, 4, status, bg=VERDE_CLA, bold=True, h='center')

# Alertas
r_al = 4 + len(indicadores) + 1
ws0.merge_cells(f'A{r_al}:D{r_al}')
c = ws0[f'A{r_al}']; c.value = "⚠  PONTOS DE ATENÇÃO"
c.fill = fill(AMARELO); c.font = Font(bold=True, size=10, color="7D6608", name='Calibri')
c.alignment = Alignment(horizontal='left', vertical='center'); c.border = bord()
ws0.row_dimensions[r_al].height = 18

alertas = [
    "R$ 4,43M da infraestrutura financiados pelas vendas → exige lançamento forte",
    "Velocidade de vendas não testada (mercado sem loteamento há +1 ano) → fazer pré-cadastro",
    "Impostos precisam de validação com contador especializado em loteamento",
    "Contrato de parceria ainda não assinado → formalizar antes de qualquer custo",
]
for j, a in enumerate(alertas):
    r2 = r_al + 1 + j
    ws0.merge_cells(f'A{r2}:D{r2}')
    c = ws0.cell(row=r2, column=1, value=f"  • {a}")
    c.fill = fill(AMARELO); c.font = Font(size=9, italic=True, color="7D6608", name='Calibri')
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    c.border = bord(); ws0.row_dimensions[r2].height = 18

r_go = r_al + len(alertas) + 2
ws0.merge_cells(f'A{r_go}:D{r_go}')
c = ws0[f'A{r_go}']; c.value = "✅  RECOMENDAÇÃO: GO — EMPREENDIMENTO VIÁVEL PARA AVANÇAR"
c.fill = fill(VERDE_CLA); c.font = Font(bold=True, size=12, color=VERDE_ESC, name='Calibri')
c.alignment = Alignment(horizontal='center', vertical='center'); c.border = bord()
ws0.row_dimensions[r_go].height = 24

path = "/Users/leonardoforx/Downloads/Papagaios/Viabilidade_Loteamento_Papagaios.xlsx"
wb.save(path)
print(f"✅ Planilha principal salva: {path}")
